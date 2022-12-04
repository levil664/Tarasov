from openpyxl import Workbook
from openpyxl.styles import Border, Side, Alignment, Font
import csv
import math
import string


class Vacancy:
    currency_to_rub = {"AZN": 35.68, "BYR": 23.91, "EUR": 59.90,
                       "GEL": 21.74, "KGS": 0.76, "KZT": 0.13, "RUR": 1,
                       "UAH": 1.64, "USD": 60.66, "UZS": 0.0055}

    def __init__(self, row_dict: dict):
        self.__name = row_dict['name']
        self.__area_name = row_dict['area_name']
        self.__publish_time = int(row_dict['published_at'][:4])
        self.__salary_from = float(row_dict['salary_from'])
        self.__salary_to = float(row_dict['salary_to'])
        self.__salary_curr = row_dict['salary_currency']
        self.__average_salary = (self.__salary_from + self.__salary_to) / 2 * self.currency_to_rub[self.__salary_curr]

    @property
    def get_name(self):
        return self.__name

    @property
    def get_area_name(self):
        return self.__area_name

    @property
    def get_publish_time(self):
        return self.__publish_time

    @property
    def get_average_salary(self):
        return self.__average_salary


class Statistic:
    def __init__(self, get_selected_vacancy: str):
        self.__selected_vacancy = get_selected_vacancy
        self.__vacancies_count = 0
        self.__cities = dict()
        self.__publish_times = dict()

        self.__salary_dynamic = dict()
        self.__vacancies_dynamic = dict()
        self.__selected_salary_dynamic = dict()
        self.__selected_vacancies_dynamic = dict()
        self.__city_salary_dynamic = dict()
        self.__city_vacancies_dynamic = dict()

        self.fulfillment = False

    @property
    def get_salary_dynamic(self):
        if not self.fulfillment:
            self.handle_information()
        return self.__salary_dynamic

    @property
    def get_vacancies_dynamic(self):
        if not self.fulfillment:
            self.handle_information()
        return self.__vacancies_dynamic

    @property
    def get_selected_salary_dynamic(self):
        if not self.fulfillment:
            self.handle_information()
        return self.__selected_salary_dynamic

    @property
    def get_selected_vacancies_dynamic(self):
        if not self.fulfillment:
            self.handle_information()
        return self.__selected_vacancies_dynamic

    @property
    def get_city_salary_dynamic(self):
        if not self.fulfillment:
            self.handle_information()
        return self.__city_salary_dynamic

    @property
    def get_city_vacancies_dynamic(self):
        if not self.fulfillment:
            self.handle_information()
        return self.__city_vacancies_dynamic

    @property
    def get_selected_vacancy(self):
        return self.__selected_vacancy

    def enter_static_data(self, data):
        for row_dict in data:
            self.update(row_dict)

    def handle_information(self):
        for publish_time in self.__publish_times.values():
            self.__salary_dynamic[publish_time.get_name] = math.floor(publish_time.get_average_salary)
            self.__vacancies_dynamic[publish_time.get_name] = publish_time.get_vacancy_count
            self.__selected_salary_dynamic[publish_time.get_name] = math.floor(publish_time.get_selected_vacancy_average_salary)
            self.__selected_vacancies_dynamic[publish_time.get_name] = publish_time.get_selected_vacancy_count

        self.__cities = dict(filter(lambda x: x[1].get_vacancy_count >= (self.__vacancies_count / 100),
                                    self.__cities.items()))
        self.__city_salary_dynamic = dict(sorted(self.__cities.items(),
                                                  key=lambda x: x[1].get_average_salary, reverse=True)[:10])
        self.__city_salary_dynamic = {key: math.floor(value.get_average_salary)
                                       for key, value in self.__city_salary_dynamic.items()}
        self.__city_vacancies_dynamic = dict(sorted(self.__cities.items(),
                                                         key=lambda x: x[1].get_vacancy_count, reverse=True)[:10])
        self.__city_vacancies_dynamic = {key: round(value.get_vacancy_count / self.__vacancies_count, 4)
                                              for key, value in self.__city_vacancies_dynamic.items()}
        self.fulfillment = True

    def update(self, row_dict: dict):
        vacancy = Vacancy(row_dict)
        if vacancy.get_area_name not in self.__cities.keys():
            self.__cities[vacancy.get_area_name] = City(vacancy)
        else:
            self.__cities[vacancy.get_area_name].update(vacancy)
        if vacancy.get_publish_time not in self.__publish_times.keys():
            self.__publish_times[vacancy.get_publish_time] = Year(vacancy, self.__selected_vacancy)
        else:
            self.__publish_times[vacancy.get_publish_time].update(vacancy)
        self.__vacancies_count += 1

    def print_statistics(self):
        self.handle_information()
        print("Динамика уровня зарплат по годам:", self.__salary_dynamic)
        print("Динамика количества вакансий по годам:", self.__vacancies_dynamic)
        print("Динамика уровня зарплат по годам для выбранной профессии:", self.__selected_salary_dynamic)
        print("Динамика количества вакансий по годам для выбранной профессии:", self.__selected_vacancies_dynamic)
        print("Уровень зарплат по городам (в порядке убывания):", self.__city_salary_dynamic)
        print("Доля вакансий по городам (в порядке убывания):", self.__city_vacancies_dynamic)


class City:
    def __init__(self, vacancy: Vacancy):
        self.__name = vacancy.get_area_name
        self.__vacancy_count = 1
        self.__all_salary = vacancy.get_average_salary
        self.__average_salary = vacancy.get_average_salary

    @property
    def get_average_salary(self):
        return self.__average_salary

    @property
    def get_vacancy_count(self):
        return self.__vacancy_count

    def update(self, vacancy: Vacancy):
        self.__vacancy_count += 1
        self.__all_salary += vacancy.get_average_salary
        self.__average_salary = self.__all_salary / self.__vacancy_count


class Year:
    def __init__(self, vacancy: Vacancy, get_selected_vacancy: str):
        self.__name = vacancy.get_publish_time
        self.__vacancy_count = 1
        self.__all_salary = vacancy.get_average_salary
        self.__average_salary = vacancy.get_average_salary
        self.__selected_vacancy = get_selected_vacancy

        self.__selected_vacancy_count = 1 if get_selected_vacancy in vacancy.get_name else 0
        self.__selected_vacancy_all_salary = vacancy.get_average_salary if get_selected_vacancy in vacancy.get_name else 0
        self.__selected_vacancy_average_salary = vacancy.get_average_salary if get_selected_vacancy in vacancy.get_name else 0

    @property
    def get_name(self):
        return self.__name

    @property
    def get_average_salary(self):
        return self.__average_salary

    @property
    def get_vacancy_count(self):
        return self.__vacancy_count

    @property
    def get_selected_vacancy_count(self):
        return self.__selected_vacancy_count

    @property
    def get_selected_vacancy_average_salary(self):
        return self.__selected_vacancy_average_salary

    def update(self, vacancy: Vacancy):
        self.__vacancy_count += 1
        self.__all_salary += vacancy.get_average_salary
        self.__average_salary = self.__all_salary / self.__vacancy_count

        if self.__selected_vacancy in vacancy.get_name:
            self.__selected_vacancy_count += 1
            self.__selected_vacancy_all_salary += vacancy.get_average_salary
            self.__selected_vacancy_average_salary = self.__selected_vacancy_all_salary / self.__selected_vacancy_count


class DataSet:
    def __init__(self, file_name: str, get_selected_vacancy: str):
        file = open(file_name, 'r', encoding='utf-8-sig')
        self.data = csv.reader(file, delimiter=',')
        self.titles = next(self.data)
        self.get_dict_row()
        self.statistic = Statistic(get_selected_vacancy)
        self.statistic.enter_static_data(self.data)

    def get_dict_row(self):
        self.data = filter(lambda row: len(row) == len(self.titles) and "" not in row, self.data)
        self.data = (dict(zip(self.titles, row)) for row in self.data)


class Report:
    title_font = Font(name='Calibri', size=11, bold=True)

    border = Border(left=Side(border_style="thin", color='FF000000'),
                    right=Side(border_style="thin", color='FF000000'),
                    top=Side(border_style="thin", color='FF000000'),
                    bottom=Side(border_style="thin", color='FF000000'))

    def __init__(self, statistic: Statistic):
        self.__statistic = statistic
        self.__book = Workbook()
        self.__publish_time_list = self.__book.active
        self.__publish_time_list.title = "Статистика по годам"
        self.__city_list = self.__book.create_sheet("Статистика по городам")

    def generate_excel(self):
        self.print_columns(['Год', 'Средняя зарплата', 'Средняя зарплата - ' + self.__statistic.get_selected_vacancy,
                            'Количество вакансий', 'Количество вакансий - ' + self.__statistic.get_selected_vacancy],
                           (list(self.__statistic.get_salary_dynamic.keys()), 'right', False),
                           (list(self.__statistic.get_salary_dynamic.values()), 'right', False),
                           (list(self.__statistic.get_selected_salary_dynamic.values()), 'right', False),
                           (list(self.__statistic.get_vacancies_dynamic.values()), 'right', False),
                           (list(self.__statistic.get_selected_vacancies_dynamic.values()), 'right', False))
        self.__book.active = self.__city_list
        self.print_columns(['Город', 'Уровень зарплат', '', 'Город', 'Доля вакансий'],
                           (list(self.__statistic.get_city_salary_dynamic.keys()), 'left', False),
                           (list(self.__statistic.get_city_salary_dynamic.values()), 'right', False),
                           (list(), 'right', False),
                           (list(self.__statistic.get_city_vacancies_dynamic.keys()), 'left', False),
                           (list(self.__statistic.get_city_vacancies_dynamic.values()), 'right', True))
        self.__book.active = self.__publish_time_list
        self.__book.save("report.xlsx")

    def fill_titles(self, titles: list):
        columns = string.ascii_uppercase[:len(titles)]
        ws = self.__book.active
        for i, column in enumerate(columns):
            if titles[i] == "":
                ws.column_dimensions[column].width = 2
                continue
            ws[column + '1'] = titles[i]
            ws[column + '1'].font = self.title_font
            ws[column + '1'].border = self.border
            ws.column_dimensions[column].width = len(titles[i]) + 2

    def print_columns(self, titles: list, *args):
        self.fill_titles(titles)
        columns = string.ascii_uppercase[:len(titles)]
        for i, arg in enumerate(args):
            self.print_column(columns[i], arg[0], arg[1], arg[2])

    def print_column(self, column_name: str, column_data: list, alignment: str, is_percent: bool):
        ws = self.__book.active
        for i in range(len(column_data)):
            ws[column_name + str(i + 2)] = str(round(column_data[i] * 100, 2)) + '%' if is_percent else column_data[i]
            if ws.column_dimensions[column_name].width < len(str(column_data[i])) + 2:
                ws.column_dimensions[column_name].width = len(str(column_data[i])) + 2
            ws[column_name + str(i + 2)].alignment = Alignment(horizontal=alignment)
            ws[column_name + str(i + 2)].border = self.border


def process_file():
    file_name = input("Введите название файла: ")
    profession_name = input("Введите название профессии: ")
    data_set = DataSet(file_name, profession_name)
    data_set.statistic.print_statistics()
    report = Report(data_set.statistic)
    report.generate_excel()


process_file()
