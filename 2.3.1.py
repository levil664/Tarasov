from jinja2 import Environment, FileSystemLoader
import doctest
import pdfkit
import numpy as np
import matplotlib.pyplot as table
from openpyxl import Workbook
from openpyxl.styles import Border, Side, Alignment, Font
import csv
import math
import string


class Vacancy:
    """
    Класс для структурирования полученных данных выбранной вакансии

    Attributes:
        __name (str) : Название вакансии
        __area_name (str) : Название города
        __publish_time (int) : Время публикации
        __salary_from (float) : Нижняя граница вилки оклада
        __salary_to (float) : Верхняя граница вилки оклада
        __salary_curr (str) : Валюта оклада
        __average_salary (float) : Среднее значение вилки оклада уже в валюте
        currency_to_rub (dict) : Курс валют к рублю
    """
    currency_to_rub = {"AZN": 35.68, "BYR": 23.91, "EUR": 59.90,
                       "GEL": 21.74, "KGS": 0.76, "KZT": 0.13, "RUR": 1,
                       "UAH": 1.64, "USD": 60.66, "UZS": 0.0055}

    def __init__(self, row_dict: dict):
        """
        Инициализирует объект Vacancy, присваивает значения из row_dict для более удобной работы с данными

        Args:
            row_dict (dict) : Обработанные данные собраные в словаре
        """
        self.__name = row_dict['name']
        self.__area_name = row_dict['area_name']
        self.__publish_time = int(row_dict['published_at'][:4])
        self.__salary_from = float(row_dict['salary_from'])
        self.__salary_to = float(row_dict['salary_to'])
        self.__salary_curr = row_dict['salary_currency']
        self.__average_salary = (self.__salary_from + self.__salary_to) / 2 * self.currency_to_rub[self.__salary_curr]

    @property
    def get_name(self):
        """
        Геттер, который возвращает название вакансии

        Returns:
            str: Название вакансии
        """
        return self.__name

    @property
    def get_area_name(self):
        """
        Геттер, который возвращает название города

        Returns:
            str: Название города
        """
        return self.__area_name

    @property
    def get_publish_time(self):
        """
        Геттер, который возвращает время публикации

        Returns:
            int: Время публикации
        """
        return self.__publish_time

    @property
    def get_average_salary(self):
        """
        Геттер, который возвращает среднее значение вилки оклада уже в валюте

        Returns:
            float: Среднее значение вилки оклада уже в валюте
        """
        return self.__average_salary

class Statistic:
    """
    Класс формирующий статистику выбранной вакансии

    Attributes:
        __selected_vacancy (str) : Название вакансии
        __vacancies_count (int) : Количество вакансий в файле
        __cities (dict) : Города и их кол-во
        __publish_times (dict) : Дата публикации вакансии
        __salary_dynamic (dict) : Динамика зарплат
        __vacancies_dynamic (dict) : Динамика количества вакансий
        __selected_salary_dynamic (dict) : Динамика зарплат вакансии
        __selected_vacancies_dynamic (dict) : Динамика количества вакансии
        __city_salary_dynamic (dict) : Динамика зарплат вакансии в выбранном городе
        __city_vacancies_dynamic (dict) : Динамика количества вакансии в выбранном городе
    """
    def __init__(self, get_selected_vacancy: str):
        """
        Инициализирует объект Statistic, получает get_selected_vacancy - выбранную вакансию

        Args:
            get_selected_vacancy (str) : Выбранная вакансия
        """
        self.__selected_vacancy = get_selected_vacancy
        self.__vacancies_count = 0
        self.__cities = dict()
        self.__publish_times = dict()

        self.__salary_dynamic = dict()
        self.__vacancies_dynamic = dict()
        self.__selected_salary_dynamic = dict()
        self.__selected_vacancies_dynamic = dict()
        self.__city_salary_dynamic = dict()
        self.__city_vacancies_dynamic = dict()

        self.fulfillment = False

    @property
    def get_salary_dynamic(self):
        """
        Геттер, который возвращает динамику зарплат

        Returns:
            dict: Динамика зарплат
        """
        if not self.fulfillment:
            self.handle_information()
        return self.__salary_dynamic

    @property
    def get_vacancies_dynamic(self):
        """
        Геттер, который возвращает динамику количества вакансий

        Returns:
            dict: Динамика количества вакансий
        """
        if not self.fulfillment:
            self.handle_information()
        return self.__vacancies_dynamic

    @property
    def get_selected_salary_dynamic(self):
        """
        Геттер, который возвращает динамику зарплат вакансии

        Returns:
            dict: Динамика зарплат вакансии
        """
        if not self.fulfillment:
            self.handle_information()
        return self.__selected_salary_dynamic

    @property
    def get_selected_vacancies_dynamic(self):
        """
        Геттер, который возвращает динамику количества вакансии

        Returns:
            dict: Динамика количества вакансии
        """
        if not self.fulfillment:
            self.handle_information()
        return self.__selected_vacancies_dynamic

    @property
    def get_city_salary_dynamic(self):
        """
        Геттер, который возвращает динамика зарплат вакансии в выбранном городе

        Returns:
            dict: динамика зарплат вакансии в выбранном городе
        """
        if not self.fulfillment:
            self.handle_information()
        return self.__city_salary_dynamic

    @property
    def get_city_vacancies_dynamic(self):
        """
        Геттер, который возвращает динамику количества вакансии в выбранном городе

        Returns:
            dict: Динамика количества вакансии в выбранном городе
        """
        if not self.fulfillment:
            self.handle_information()
        return self.__city_vacancies_dynamic

    @property
    def get_selected_vacancy(self):
        """
        Геттер, который возвращает название вакансии

        Returns:
            str: Название вакансии
        """
        return self.__selected_vacancy

    def enter_static_data(self, data):
        """
        Обновляет значение строк файла
        """
        for row_dict in data:
            self.update(row_dict)

    def handle_information(self):
        """
        Обрабатывает  по ТЗ информацию из файла, для дальнейшей работы
        """
        for publish_time in self.__publish_times.values():
            self.__salary_dynamic[publish_time.get_name] = math.floor(publish_time.get_average_salary)
            self.__vacancies_dynamic[publish_time.get_name] = publish_time.get_vacancy_count
            self.__selected_salary_dynamic[publish_time.get_name] = math.floor(publish_time.get_selected_vacancy_average_salary)
            self.__selected_vacancies_dynamic[publish_time.get_name] = publish_time.get_selected_vacancy_count

        self.__cities = dict(filter(lambda x: x[1].get_vacancy_count >= (self.__vacancies_count / 100),
                                    self.__cities.items()))
        self.__city_salary_dynamic = dict(sorted(self.__cities.items(),
                                                  key=lambda x: x[1].get_average_salary, reverse=True)[:10])
        self.__city_salary_dynamic = {key: math.floor(value.get_average_salary)
                                       for key, value in self.__city_salary_dynamic.items()}
        self.__city_vacancies_dynamic = dict(sorted(self.__cities.items(),
                                                         key=lambda x: x[1].get_vacancy_count, reverse=True)[:10])
        self.__city_vacancies_dynamic = {key: round(value.get_vacancy_count / self.__vacancies_count, 4)
                                              for key, value in self.__city_vacancies_dynamic.items()}
        self.fulfillment = True

    def update(self, row_dict: dict):
        """
        Обновляет количественные значения и время
        """
        vacancy = Vacancy(row_dict)
        if vacancy.get_area_name not in self.__cities.keys():
            self.__cities[vacancy.get_area_name] = City(vacancy)
        else:
            self.__cities[vacancy.get_area_name].update(vacancy)
        if vacancy.get_publish_time not in self.__publish_times.keys():
            self.__publish_times[vacancy.get_publish_time] = Year(vacancy, self.__selected_vacancy)
        else:
            self.__publish_times[vacancy.get_publish_time].update(vacancy)
        self.__vacancies_count += 1


    def print_statistics(self):
        """
        Фунция, которая при надобности будет печатать информацию (для проверки программы)
        """
        self.handle_information()
        print("Динамика уровня зарплат по годам:", self.__salary_dynamic)
        print("Динамика количества вакансий по годам:", self.__vacancies_dynamic)
        print("Динамика уровня зарплат по годам для выбранной профессии:", self.__selected_salary_dynamic)
        print("Динамика количества вакансий по годам для выбранной профессии:", self.__selected_vacancies_dynamic)
        print("Уровень зарплат по городам (в порядке убывания):", self.__city_salary_dynamic)
        print("Доля вакансий по городам (в порядке убывания):", self.__city_vacancies_dynamic)


class City:
    """
    Класс отвечающий за статистику, которая как-либо связана с городами

    Attributes:
        __name (str) : Название Города
        __vacancy_count (int) : Количество вакансий в этом городе
        __all_salary (float): Сумма всех средних зарплат в этом городе
        __average_salary (float) : Средняя зарплата по городу
    """
    def __init__(self, vacancy: Vacancy):
        """
            Инициализирует объект City, получает информацию о вакансии

            Args:
                vacancy (Vacancy) : инофрмация о вакансии
        """
        self.__name = vacancy.get_area_name
        self.__vacancy_count = 1
        self.__all_salary = vacancy.get_average_salary
        self.__average_salary = vacancy.get_average_salary

    @property
    def get_average_salary(self):
        """
        Геттер, который возвращает среднюю запрлату по городу

        Returns:
            float: Среднюю запрлату по городу
        """
        return self.__average_salary

    @property
    def get_vacancy_count(self):
        """
        Геттер, который возвращает количество вакансий в городе

        Returns:
            int: количество вакансий в городе
        """
        return self.__vacancy_count

    def update(self, vacancy: Vacancy):
        """
        Вычисляет среднюю зарплату по городу с каждым вызовом функции
        """
        self.__vacancy_count += 1
        self.__all_salary += vacancy.get_average_salary
        self.__average_salary = self.__all_salary / self.__vacancy_count


class Year:
    """
    Класс отвечающий за динамику выбранной вакансии по годам

    Attributes:
        __name (str) : Время публикации
        __vacancy_count (int) : Количество вакансий в файле
        __all_salary (float) : Сумма зарплат
        __average_salary (float) : Средняя зарплата
        __selected_vacancy (str) : Выбранная вакансия
        __selected_vacancy_count (int) : Количество выбранной вакансии в файле
        __selected_vacancy_all_salary (float | int) : Сумма зарплат вакансии
        __selected_vacancy_average_salary (float | int) : Средняя зарплата вакансии
    """
    def __init__(self, vacancy: Vacancy, get_selected_vacancy: str):
        """
        Инициализирует объект Year, получает get_selected_vacancy - выбранную вакансию,
        vacancy - информацию о вакансии.

        Args:
            get_selected_vacancy (str) : Выбранная вакансия
            vacancy (Vacancy) : Информацию о вакансии
        """
        self.__name = vacancy.get_publish_time
        self.__vacancy_count = 1
        self.__all_salary = vacancy.get_average_salary
        self.__average_salary = vacancy.get_average_salary
        self.__selected_vacancy = get_selected_vacancy

        self.__selected_vacancy_count = 1 if get_selected_vacancy in vacancy.get_name else 0
        self.__selected_vacancy_all_salary = vacancy.get_average_salary if get_selected_vacancy in vacancy.get_name else 0
        self.__selected_vacancy_average_salary = vacancy.get_average_salary if get_selected_vacancy in vacancy.get_name else 0

    @property
    def get_name(self):
        """
        Геттер, который возвращает время публикации

        Returns:
            dict: Время публикации
        """
        return self.__name

    @property
    def get_average_salary(self):
        """
        Геттер, который возвращает средняя зарплата

        Returns:
            dict: Средняя зарплата
        """
        return self.__average_salary

    @property
    def get_vacancy_count(self):
        """
        Геттер, который возвращает количество вакансий в файле

        Returns:
            dict: Количество вакансий в файле
        """
        return self.__vacancy_count

    @property
    def get_selected_vacancy_count(self):
        """
        Геттер, который возвращает количество выбранной вакансии в файле

        Returns:
            dict: Количество выбранной вакансии в файле
        """
        return self.__selected_vacancy_count

    @property
    def get_selected_vacancy_average_salary(self):
        """
        Геттер, который возвращает среднюю зарплату выбранной вакансии

        Returns:
            dict: Средняя зарплата выбранной вакансии
        """
        return self.__selected_vacancy_average_salary

    def update(self, vacancy: Vacancy):
        """
        Обновляет значения количество вакансий и среднюю зарплату выбранной вакансии
        """
        self.__vacancy_count += 1
        self.__all_salary += vacancy.get_average_salary
        self.__average_salary = self.__all_salary / self.__vacancy_count

        if self.__selected_vacancy in vacancy.get_name:
            self.__selected_vacancy_count += 1
            self.__selected_vacancy_all_salary += vacancy.get_average_salary
            self.__selected_vacancy_average_salary = self.__selected_vacancy_all_salary / self.__selected_vacancy_count


class DataSet:
    """
    Класс для сборки и передачи полученных данных из файла

    Attributes:
        data (_reader) : Считанный файл
        titles (list[str]) : Название каждого столбца
    """
    def __init__(self, file_name: str, get_selected_vacancy: str):
        """
        Инициализирует объект DataSet, получает значения file_name для работы с файлом
         и get_selected_vacancy для работы с выбранной вакансией

        Args:
            file_name (str) : Название csv файла
            get_selected_vacancy (str) : Название выбранной ванкансии
        """
        file = open(file_name, 'r', encoding='utf-8-sig')
        self.data = csv.reader(file, delimiter=',')
        self.titles = next(self.data)
        self.get_dict_row()
        self.statistic = Statistic(get_selected_vacancy)
        self.statistic.enter_static_data(self.data)

    def get_dict_row(self):
        """Убирает пустые поля и формирует словарь для удобной работы с данными"""
        self.data = filter(lambda row: len(row) == len(self.titles) and "" not in row, self.data)
        self.data = (dict(zip(self.titles, row)) for row in self.data)


class Report:
    """
    Класс для формирования вывода обработанной информации в соответсвии с ТЗ

    Attributes:
        __statistic (Statistic) : Конечная статистика по файлу
        __book (Workbook) : Форма для формирования вывода
        sheet_1_headers (list(str)) : Название колонок для формирования таблицы 1
        sheet_1_columns (list[list]) : Значимые поля таблицы 1
        sheet_1_rows (list[list[str]]) : Название строк для формирования таблицы 1
        sheet_2_headers (list(str)) : Название колонок для формирования таблицы 2
        sheet_2_columns (list[list]) : Значимые поля таблицы 2
        sheet_2_rows (list[list[str]]) : Название строк для формирования таблицы 2
    """
    title_font = Font(name='Calibri', size=11, bold=True)

    border = Border(left=Side(border_style="thin", color='FF000000'),
                    right=Side(border_style="thin", color='FF000000'),
                    top=Side(border_style="thin", color='FF000000'),
                    bottom=Side(border_style="thin", color='FF000000'))

    def __init__(self, statistic: Statistic):
        self.__statistic = statistic
        self.__book = Workbook()
        self.__publish_time_list = self.__book.active
        self.__publish_time_list.title = "Статистика по годам"
        self.__city_list = self.__book.create_sheet("Статистика по городам")
        self.sheet_1_headers = ["Год", "Средняя зарплата", "Средняя зарплата - " + self.__statistic.get_selected_vacancy,
                                "Количество вакансий", "Количество вакансий - " + self.__statistic.get_selected_vacancy]
        sheet_1_columns = [list(self.__statistic.get_salary_dynamic.keys()), list(self.__statistic.get_salary_dynamic.values()),
                           list(self.__statistic.get_selected_salary_dynamic.values()), list(self.__statistic.get_vacancies_dynamic.values()),
                           list(self.__statistic.get_selected_vacancies_dynamic.values())]
        self.sheet_1_rows = self.get_table_rows(sheet_1_columns)
        self.sheet_2_headers = ["Город", "Уровень зарплат", " ", "Город", "Доля вакансий"]
        sheet_2_columns = [list(self.__statistic.get_city_salary_dynamic.keys()), list(self.__statistic.get_city_salary_dynamic.values()),
                           ["" for _ in self.__statistic.get_city_salary_dynamic.keys()], list(self.__statistic.get_city_vacancies_dynamic.keys()),
                           list(map(self.get_percents, self.__statistic.get_city_vacancies_dynamic.values()))]
        self.sheet_2_rows = self.get_table_rows(sheet_2_columns)

    def get_percents(self, value):
        """
        Функция для формирования процентов

        Attributes:
            value (float) : Значение, которое надо перевести в проценты

        Returns:
            str: Переданное value в процентах
        """
        return f"{round(value * 100, 2)}%"

    def get_table_rows(self, columns: list):
        """
        Геттер, который возвращет list-строк колонок

        Attributes:
            columns (list) : Значение, у которого надо выбрать строки

        Returns:
            str: Переданное value в процентах
        """
        rows_list = [["" for _ in range(len(columns))] for _ in range(len(columns[0]))]
        for col in range(len(columns)):
            for cell in range(len(columns[col])):
                rows_list[cell][col] = columns[col][cell]
        return rows_list

    def generate_image(self):
        """
        Функция делает и показывет графики отражающие динамику зарплат вакансии
        """
        temp, matrix_of_graphs = table.subplots(2, 2)
        width = 0.4
        abscissa = np.arange(len(self.__statistic.get_salary_dynamic.keys()))
        matrix_of_graphs[0, 0].bar(abscissa-width/2, self.__statistic.get_salary_dynamic.values(), width=width, label='средняя з/п')
        matrix_of_graphs[0, 0].bar(abscissa+width/2, self.__statistic.get_selected_salary_dynamic.values(), width=width, label='з/п программист')
        matrix_of_graphs[0, 0].set_xticks(abscissa, self.__statistic.get_salary_dynamic.keys())
        matrix_of_graphs[0, 0].set_xticklabels(self.__statistic.get_salary_dynamic.keys(), rotation='vertical', va='top', ha='center')
        matrix_of_graphs[0, 0].set_title('Уровень зарплат по годам')
        matrix_of_graphs[0, 0].grid(True, axis='y')
        matrix_of_graphs[0, 0].tick_params(axis='both', labelsize=8)
        matrix_of_graphs[0, 0].legend(fontsize=8)

        matrix_of_graphs[0, 1].bar(abscissa-width/2, self.__statistic.get_vacancies_dynamic.values(), width=width, label='Количество вакансий')
        matrix_of_graphs[0, 1].bar(abscissa+width/2, self.__statistic.get_selected_vacancies_dynamic.values(), width=width, label='Количество вакансий\nпрограммист')
        matrix_of_graphs[0, 1].set_xticks(abscissa, self.__statistic.get_vacancies_dynamic.keys())
        matrix_of_graphs[0, 1].set_xticklabels(self.__statistic.get_vacancies_dynamic.keys(), rotation='vertical', va='top', ha='center')
        matrix_of_graphs[0, 1].set_title('Количество вакансий по годам')
        matrix_of_graphs[0, 1].grid(True, axis='y')
        matrix_of_graphs[0, 1].tick_params(axis='both', labelsize=8)
        matrix_of_graphs[0, 1].legend(fontsize=8)

        matrix_of_graphs[1, 0].set_title("Уровень зарплат по городам")
        matrix_of_graphs[1, 0].invert_yaxis()
        data = self.__statistic.get_city_salary_dynamic
        courses = list(data.keys())
        courses = [label.replace(' ', '\n').replace('-', '-\n') for label in courses]
        values = list(data.values())
        matrix_of_graphs[1, 0].tick_params(axis='both', labelsize=8)
        matrix_of_graphs[1, 0].set_yticklabels(courses, fontsize=6, va='center', ha='right')
        matrix_of_graphs[1, 0].barh(courses, values)
        matrix_of_graphs[1, 0].grid(True, axis='x')

        other = 1 - sum((list(self.__statistic.get_city_vacancies_dynamic.values())))
        new_dic = {'Другие': other}
        new_dic.update(self.__statistic.get_city_vacancies_dynamic)
        area_count_dic = new_dic
        labels = list(area_count_dic.keys())
        sizes = list(area_count_dic.values())
        matrix_of_graphs[1, 1].pie(sizes, labels=labels, textprops={'fontsize': 6})
        matrix_of_graphs[1, 1].axis('scaled')
        matrix_of_graphs[1, 1].set_title("Доля вакансий по городам")
        table.tight_layout()
        table.savefig('graph.png', dpi=300)
        table.show()

    def generate_pdf(self, choice: str):
        """
        Функция генерирует пдф в соответсвии шаблону html

        Attributes:
            choice (str) : Значение, которое хочет пользователь видеть в пдф только графики/только таблицу/все сразу
        """
        image_file = 'graph.png'
        env = Environment(loader=FileSystemLoader('.'))
        if (choice == "Вакансии"):
            template = env.get_template("pdf_template_img.html")
            pdf_template = template.render({"title": "Аналитика по зарплатам и городам для профессии " + self.__statistic.get_selected_vacancy,
                                            "image_file": image_file,
                                            })
        elif (choice == "Статистика"):
            template = env.get_template("pdf_template_statistic.html")
            pdf_template = template.render(
                {"title": "Аналитика по зарплатам и городам для профессии " + self.__statistic.get_selected_vacancy,
                 "years_title": "Статистика по годам",
                 "years_headers": self.sheet_1_headers,
                 "years_rows": self.sheet_1_rows,
                 "cities_title": "Статистика по городам",
                 "cities_headers": self.sheet_2_headers,
                 "count_columns": len(self.sheet_2_headers),
                 "cities_rows": self.sheet_2_rows
                 })
        else:
            template = env.get_template("pdf_template.html")
            pdf_template = template.render(
                {"title": "Аналитика по зарплатам и городам для профессии " + self.__statistic.get_selected_vacancy,
                 "image_file": image_file,
                 "years_title": "Статистика по годам",
                 "years_headers": self.sheet_1_headers,
                 "years_rows": self.sheet_1_rows,
                 "cities_title": "Статистика по городам",
                 "cities_headers": self.sheet_2_headers,
                 "count_columns": len(self.sheet_2_headers),
                 "cities_rows": self.sheet_2_rows
                 })
        config = pdfkit.configuration(wkhtmltopdf=r"C:/Program Files/wkhtmltopdf/bin/wkhtmltopdf.exe")
        pdfkit.from_string(pdf_template, 'report.pdf', configuration=config, options={"enable-local-file-access": True})

    def generate_excel(self):
        """
        Функция генерирует таблицу excel динамики зарплат вакансии
        """
        self.print_columns(['Год', 'Средняя зарплата',
                            'Средняя зарплата - ' + self.__statistic.get_selected_vacancy,
                            'Количество вакансий', 'Количество вакансий - ' + self.__statistic.get_selected_vacancy],
                           (list(self.__statistic.get_salary_dynamic.keys()), 'right', False),
                           (list(self.__statistic.get_salary_dynamic.values()), 'right', False),
                           (list(self.__statistic.get_selected_salary_dynamic.values()), 'right', False),
                           (list(self.__statistic.get_vacancies_dynamic.values()), 'right', False),
                           (list(self.__statistic.get_selected_vacancies_dynamic.values()), 'right', False))
        self.__book.active = self.__city_list
        self.print_columns(['Город', 'Уровень зарплат', '', 'Город', 'Доля вакансий'],
                           (list(self.__statistic.get_city_salary_dynamic.keys()), 'left', False),
                           (list(self.__statistic.get_city_salary_dynamic.values()), 'right', False),
                           (list(), 'right', False),
                           (list(self.__statistic.get_city_vacancies_dynamic.keys()), 'left', False),
                           (list(self.__statistic.get_city_vacancies_dynamic.values()), 'right', True))
        self.__book.active = self.__publish_time_list
        self.__book.save("report.xlsx")

    def fill_titles(self, titles: list):
        """
        Заполняет заголовки в таблице excel

        Attributes:
            titles (list) : Названия столбцов
        """
        columns = string.ascii_uppercase[:len(titles)]
        ws = self.__book.active
        for i, column in enumerate(columns):
            if titles[i] == "":
                ws.column_dimensions[column].width = 2
                continue
            ws[column + '1'] = titles[i]
            ws[column + '1'].font = self.title_font
            ws[column + '1'].border = self.border
            ws.column_dimensions[column].width = len(titles[i]) + 2

    def print_columns(self, titles: list, *args):
        """
        Заполняет столбцы в таблице excel

        Attributes:
            titles (list) : Названия столбцов
            *args (Any) : Аргументы столбцов
        """
        self.fill_titles(titles)
        columns = string.ascii_uppercase[:len(titles)]
        for i, arg in enumerate(args):
            self.print_column(columns[i], arg[0], arg[1], arg[2])

    def print_column(self, column_name: str, column_data: list, alignment: str, is_percent: bool):
        """
        Заполняет таблицу excel

        Attributes:
            column_name (str) : Заголовки столбцов
            column_data (list) : Данные столбцов
            alignment (str) : Выравнивание по горизонтали
            is_percent (bool) : В процентах значение или нет
        """
        ws = self.__book.active
        for i in range(len(column_data)):
            ws[column_name + str(i + 2)] = str(round(column_data[i] * 100, 2)) + '%' if is_percent else column_data[i]
            if ws.column_dimensions[column_name].width < len(str(column_data[i])) + 2:
                ws.column_dimensions[column_name].width = len(str(column_data[i])) + 2
            ws[column_name + str(i + 2)].alignment = Alignment(horizontal=alignment)
            ws[column_name + str(i + 2)].border = self.border


def final_process():
    """
    Ввод данных пользователя и передача их в классы
    """
    file_name = 'vacancies_by_year.csv'
    profession_name = 'Программист'
    # file_name = input("Введите название файла: ")
    # profession_name = input("Введите название профессии: ")
    data_set = DataSet(file_name, profession_name)
    # data_set.statistic.print_statistics()
    report = Report(data_set.statistic)
    report.generate_excel()
    report.generate_image()
    report.generate_pdf(input('Введите данные для печати: '))

final_process()